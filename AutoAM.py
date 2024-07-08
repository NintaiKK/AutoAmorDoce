from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from tkinter import messagebox
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl import load_workbook

class AutoAmorDoceApp:
    
    def __init__(self, master):
        
        self.master = master
        master.title('AutoAmorDoce')
        master.geometry('800x500')
        master.resizable(0, 0)

        self.frame_a = tk.Frame(master)
        self.frame_b = tk.Frame(master)
        self.frame_c = tk.Frame(master)

        self.lbl_vrd = tk.Label(
            self.frame_a,
            height=30, width=45,
            text='AutoAmorDoce',
            bg='pink')
        
        self.lbl_vrd.pack(side=tk.TOP)

        self.lbl_coiso = tk.Label(
            self.frame_a,
            height=3,
            width=45,
            text='Desenvolvido com amor por: PakiPakecitus Pêssego',
            bg='pink')
        
        self.lbl_coiso.pack(side=tk.TOP)

        self.lbl_spcs = tk.Label(
            self.frame_c,
            width=25)
        
        self.lbl_spcs.pack()

        self.btn_cadastro = tk.Button(
            self.frame_b,
            height=3,
            width=15,
            text='Cadastrar conta',
            command=self.jan_cadastro)
        
        self.btn_cadastro.pack()

        lbl_vzo = tk.Label(
            self.frame_b,
            height = 1)

        lbl_vzo.pack()

        self.btn_auto = tk.Button(
            self.frame_b,
            height=3,
            width=15,
            text='Selecionar conta',
            command=self.janAuto)
        
        self.btn_auto.pack()

        self.frame_a.pack(side=tk.LEFT)
        self.frame_c.pack(side=tk.LEFT)
        self.frame_b.pack(side=tk.LEFT)

    def jan_cadastro(self):
        self.cadastro_window = tk.Toplevel(self.master)
        self.cadastro_window.title('Cadastrar conta')
        self.cadastro_window.geometry('500x300')

        lbl_login = tk.Label(
            self.cadastro_window,
            text='Login')
        
        lbl_login.pack()

        self.entry_login = tk.Entry(
            self.cadastro_window,
            width=50)
        
        self.entry_login.pack()

        lbl_senha = tk.Label(
            self.cadastro_window,
            text='Senha')
        
        lbl_senha.pack()

        self.senha_entry = tk.Entry(
            self.cadastro_window,
            width=50)
        
        self.senha_entry.pack()

        lbl_apelido = tk.Label(
            self.cadastro_window,
            text='Apelido')
        
        lbl_apelido.pack()

        self.entry_apelido = tk.Entry(
            self.cadastro_window,
            width=50)
        
        self.entry_apelido.pack()

        lbl_fts = tk.Label(
            self.cadastro_window,
            height=1)
        
        lbl_fts.pack()

        btn_ok = tk.Button(
            self.cadastro_window,
            text='OK',
            width=5,
            height=1,
            command=self.cadastro_conta)
        
        btn_ok.pack()

    def janAuto(self):
        
        self.auto_window = tk.Toplevel(self.master)
        self.auto_window.title('Cadastrar conta')
        self.auto_window.geometry('500x300')

        lbl_text = tk.Label(
            self.auto_window,
            text = 'Selecione a conta')
        
        lbl_text.pack()

        self.combobox_linhas = ttk.Combobox(
            self.auto_window,
            width=10)
        
        self.combobox_linhas.pack()

        try:
            wb = load_workbook('autoAM.xlsx')
        except FileNotFoundError:
            return

        planilha = wb.active

        linhas_preenchidas = []

        for row in range(1, 16):
            last_col_value = planilha.cell(row=row, column=3).value
            if last_col_value is not None:
                linhas_preenchidas.append(last_col_value)

        self.combobox_linhas['values'] = linhas_preenchidas
        

    def cadastro_conta(self):
        
        try:
            wb = load_workbook('autoAM.xlsx')
        except FileNotFoundError:
            wb = Workbook()

        planilha = wb.active

        get_login = self.entry_login.get()
        get_senha = self.senha_entry.get()
        get_apelido = self.entry_apelido.get()

        planilha.append([get_login, get_senha, get_apelido])

        wb.save('autoAM.xlsx')

        self.entry_login.delete(0, tk.END)
        self.senha_entry.delete(0, tk.END)
        self.entry_apelido.delete(0, tk.END)

    def carregar_linhas_preenchidas(self):
        try:
            wb = load_workbook('autoAM.xlsx')
        except PermissionError:
            messagebox.showerror("Erro", "Banco de dados 'autoAM.xlsx' aberto.")
            return

        planilha = wb.active

        linhas_preenchidas = []

        for col in range(1, 4):
            for row in range(1, 16):
                cell_value = planilha.cell(row=row, column=col).value
                if cell_value is not None:
                    linhas_preenchidas.append(f'{row}-{col}')

        self.combobox_linhas['values'] = linhas_preenchidas

def main():
    janela = tk.Tk()
    app = AutoAmorDoceApp(janela)
    janela.mainloop()

if __name__ == "__main__":
    main()
